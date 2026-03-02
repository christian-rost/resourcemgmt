"""Excel generation service using openpyxl and the XQT5 template."""
import calendar
import io
import logging
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

logger = logging.getLogger(__name__)

TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent.parent
    / "docs"
    / "XQT5_Leistungsnachweis_mit_Ort_Vorlage_v2_0.xltx"
)

DATA_START_ROW = 14
DATA_END_ROW = 44
SUM_ROW = 45
SUM_COL = 8  # column H


def generate_timesheet_excel(
    entries: list[dict],
    project: dict,
    year: int,
    month: int,
    config: dict,
    user_display: str = "",
) -> bytes:
    """Generate an Excel timesheet based on the XQT5 Leistungsnachweis template."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    wb.template = False  # Speichern als .xlsx, nicht als .xltx-Template
    ws = wb["XQT5-LN"]

    customer = project.get("customers") or {}
    customer_name = customer.get("name", "")
    project_name = project.get("name", "")
    project_short = project.get("short_code", "")
    proj_label = f"{project_short} – {project_name}" if project_short else project_name

    # --- Header ---
    ws["D7"] = user_display
    ws["D8"] = customer_name
    ws["D9"] = proj_label
    ws["D10"] = date(year, month, 1)

    # --- Aggregate entries by date ---
    entry_by_date: dict[str, dict] = {}
    for e in entries:
        d_str = e["entry_date"]
        if d_str not in entry_by_date:
            entry_by_date[d_str] = {"hours": 0.0, "break_hours": 0.0, "comments": [], "start_time": None, "end_time": None}
        entry_by_date[d_str]["hours"] += float(e.get("hours") or 0)
        entry_by_date[d_str]["break_hours"] += float(e.get("break_hours") or 0)
        if e.get("comment"):
            entry_by_date[d_str]["comments"].append(e["comment"])
        # start_time = min, end_time = max across all entries of the day
        st = e.get("start_time")
        et = e.get("end_time")
        if st:
            if entry_by_date[d_str]["start_time"] is None or st < entry_by_date[d_str]["start_time"]:
                entry_by_date[d_str]["start_time"] = st
        if et:
            if entry_by_date[d_str]["end_time"] is None or et > entry_by_date[d_str]["end_time"]:
                entry_by_date[d_str]["end_time"] = et

    # --- Clear all template data rows ---
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        for col in range(1, 11):  # A–J
            ws.cell(row=row, column=col).value = None

    # --- Fill calendar days ---
    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, days_in_month + 1):
        row = DATA_START_ROW + day - 1
        d = date(year, month, day)
        d_str = d.isoformat()

        ws.cell(row=row, column=2).value = d  # B: Datum

        if d_str in entry_by_date:
            ed = entry_by_date[d_str]
            # Newlines und mehrfache Leerzeichen aus Kommentaren entfernen
            cleaned = [" ".join(c.split()) for c in ed["comments"] if c.strip()]
            comment = "; ".join(cleaned)
            taetigkeit = ws.cell(row=row, column=4)
            taetigkeit.value = comment  # D: Tätigkeit
            taetigkeit.alignment = Alignment(wrap_text=True, vertical="top")
            if ed["start_time"]:
                cell = ws.cell(row=row, column=5)  # E: Arbeitsbeginn
                h, m = int(ed["start_time"][:2]), int(ed["start_time"][3:5])
                cell.value = (h * 60 + m) / 1440
                cell.number_format = "HH:MM"
            if ed["end_time"]:
                cell = ws.cell(row=row, column=6)  # F: Arbeitsende
                h, m = int(ed["end_time"][:2]), int(ed["end_time"][3:5])
                cell.value = (h * 60 + m) / 1440
                cell.number_format = "HH:MM"
            if ed["break_hours"]:
                pause_cell = ws.cell(row=row, column=7)
                pause_cell.value = ed["break_hours"] / 24  # Excel-Zeit: Bruchteil eines Tages
                pause_cell.number_format = "HH:MM"
            ws.cell(row=row, column=8).value = ed["hours"]  # H: Gesamt [h]

    # --- Update SUM formula to match actual month length ---
    last_data_row = DATA_START_ROW + days_in_month - 1
    ws.cell(row=SUM_ROW, column=SUM_COL).value = f"=SUM(H{DATA_START_ROW}:H{last_data_row})"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
