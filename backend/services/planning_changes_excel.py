"""Excel-Report für Planungsänderungen (openpyxl, Corporate Design XQT5)."""
import io
import logging
from datetime import datetime, date

import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side
)

logger = logging.getLogger(__name__)

COLOR_PRIMARY = "EE7F00"   # Orange
COLOR_DARK    = "213452"   # Dunkelblau
COLOR_LIGHT   = "F5F5F5"
COLOR_WHITE   = "FFFFFF"

ACTION_LABELS = {
    "create": "Erstellt",
    "update": "Geändert",
    "delete": "Gelöscht",
}

MONTHS_DE = {
    1: "Januar", 2: "Februar", 3: "März", 4: "April",
    5: "Mai", 6: "Juni", 7: "Juli", 8: "August",
    9: "September", 10: "Oktober", 11: "November", 12: "Dezember",
}


def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def generate_changes_excel(
    changes: list,
    from_date: str,
    to_date: str,
    config: dict,
    users_by_id: dict,
    projects_by_id: dict,
) -> bytes:
    """Erzeugt Excel-Report der Planungsänderungen. Gibt Bytes zurück."""
    company = config.get("company_name", "XQT5 Ressource")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planungsänderungen"

    # ── Spaltenbreiten ────────────────────────────────────────────────────────
    col_widths = [20, 12, 22, 22, 30, 16, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # ── Titel-Zeile ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"{company} – Planungsänderungen {from_date} bis {to_date}"
    title_cell.font = Font(bold=True, color=COLOR_WHITE, size=13)
    title_cell.fill = _header_fill(COLOR_DARK)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Header-Zeile ─────────────────────────────────────────────────────────
    headers = [
        "Zeitstempel", "Aktion", "Geändert von",
        "Betroffener Berater", "Projekt", "Zeitraum",
        "Alt (Std)", "Neu (Std)",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color=COLOR_WHITE)
        cell.fill = _header_fill(COLOR_PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[2].height = 22

    # ── Datenzeilen ───────────────────────────────────────────────────────────
    row_idx = 3
    for i, change in enumerate(changes):
        # Timestamp
        ts_raw = change.get("changed_at", "")
        try:
            ts = datetime.fromisoformat(ts_raw.replace("Z", "+00:00"))
            ts_str = ts.strftime("%d.%m.%Y %H:%M")
        except Exception:
            ts_str = ts_raw

        action = ACTION_LABELS.get(change.get("action", ""), change.get("action", ""))

        changer_id = change.get("changed_by") or ""
        changer = users_by_id.get(changer_id, {})
        changer_name = changer.get("display_name") or changer.get("username") or changer_id

        affected_id = change.get("affected_user_id") or ""
        affected = users_by_id.get(affected_id, {})
        affected_name = affected.get("display_name") or affected.get("username") or affected_id

        proj_id = change.get("project_id") or ""
        project = projects_by_id.get(proj_id, {})
        proj_name = project.get("name") or proj_id
        cust = (project.get("customers") or {}).get("name") or ""
        proj_label = f"{cust} / {proj_name}" if cust else proj_name

        yr = change.get("plan_year")
        mo = change.get("plan_month")
        zeitraum = f"{MONTHS_DE.get(mo, mo)} {yr}" if yr and mo else ""

        old_hours = ""
        new_hours = ""
        old_d = change.get("old_data") or {}
        new_d = change.get("new_data") or {}
        if old_d.get("hours") is not None:
            old_hours = str(old_d["hours"])
        if new_d.get("hours") is not None:
            new_hours = str(new_d["hours"])

        row_data = [ts_str, action, changer_name, affected_name, proj_label, zeitraum, old_hours, new_hours]
        fill = _header_fill(COLOR_LIGHT) if i % 2 == 0 else None

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fill:
                cell.fill = fill
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="center")

        row_idx += 1

    # ── Zähler-Zeile ─────────────────────────────────────────────────────────
    ws.cell(row=row_idx + 1, column=1, value=f"Gesamt: {len(changes)} Einträge").font = Font(italic=True, color="888888")

    # ── Freeze header ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
